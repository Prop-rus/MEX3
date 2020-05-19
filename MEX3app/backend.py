import zipfile
from datetime import datetime
from .models import tmp_dir_ready_status
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from MEX3.settings import MEDIA_ROOT, BASE_DIR
import random
import os
from patoolib import extract_archive


def getFileExt(filename):
    ''' checking file format'''  # TODO its better to rewrite in re
    if filename.find('.zip') != -1:
        return '.zip'
    elif filename.find('.rar') != -1:
        return '.rar'
    elif filename.find('.7z') != -1:
        return '.7z'
    elif filename.find('.xls') != -1:
        return '.xls'
    elif filename.find('.xlsx') != -1:
        return '.xlsx'
    else:
        return 'wrong format'


def handle_uploaded_file(files, token, dirpath=BASE_DIR):
    ''' Creates new tmp dir and uploads files there, returning dir path and file extension'''

    if dirpath == BASE_DIR:
        tmp_dir_ready_status.objects.filter(token=token).delete()  # need to delete old object to exclude ajax mistakes
        pat = 'MEX3app/filestoobr/'
        name_dir = str(random.randint(1, 1000)) + datetime.now().strftime('%Y%m%d%H%M%S')
        new_dir = pat + name_dir
        target_dir = os.path.join(new_dir)
        os.mkdir(path=target_dir)
        save_in_db = tmp_dir_ready_status.objects.create(token=token, dir_name=name_dir,
                                                         res_ready=False)  # table that mapping the status of executing to say ajax when itsready
    else:
        new_dir = dirpath
        target_dir = os.path.join(new_dir)
        name_dir = ''
        save_in_db = tmp_dir_ready_status.objects.create(token=token, dir_name=new_dir, res_ready=False)
    quantfiles = len(files)
    ext = ''
    if quantfiles == 1:
        ext = getFileExt(str(files[0]))
        if ext == 'wrong format':  # files should be either archives either excel
            return '', '', '', '', True

    for f in files:
        nfile = str(f)
        ext = getFileExt(nfile)
        if quantfiles > 1 and ext != '.xls' and ext != '.xlsx':  # if files are multiplies the should be in single format
            return '', '', '', '', True
        with open(target_dir + '/' + nfile, 'wb+') as dest:  # downloading files
            for chunk in f.chunks():
                dest.write(chunk)
    new_files = files
    for loaded_file in os.listdir(target_dir):
        if getFileExt(loaded_file) in ['.7z', '.rar', '.zip']:
            before_files = set(os.listdir(target_dir))
            extract_archive(target_dir + '/' + loaded_file, verbosity=0, outdir=os.path.join(new_dir))
            os.remove(target_dir + '/' + loaded_file)
            after_files = set(os.listdir(target_dir))
            new_files = list(after_files - before_files)  # using in form1

    return target_dir, name_dir, new_files, save_in_db, False


def form0Unit(dir_path, cellcomm, cellvalue):
    '''Combines uploaded files in one dict, then writes it in result file'''

    filelist = os.listdir(dir_path)
    res_wb = Workbook()  # creating a workbook to write result registry
    res_wsht = res_wb.active
    res_wsht.title = 'Лист с результатами'
    for file in filelist:  # opening files source files and searching needed values
        wrkbook = load_workbook(dir_path + '/' + file, read_only=True)
        wrksheet = wrkbook.active
        comment = wrksheet.cell(column=int(cellcomm[0]), row=int(cellcomm[1])).value
        value = wrksheet.cell(column=int(cellvalue[0]), row=int(cellvalue[1])).value
        list_to_add = []
        list_to_add.append(file)
        list_to_add.append(comment)
        list_to_add.append(value)
        res_wsht.append(list_to_add)
        wrkbook.close()
    os.mkdir(dir_path + '/result')
    res_wb.save(dir_path + '/result/' + 'result.xlsx')  # saving a result regisrty
    res_wb.close()


def form1Unit(dir_path,
              files_reg, files_tofill, cellcomm, cellvalue, reg_com,
              reg_value):
    '''opens the regisrty and copies values to template files by rows'''

    target_dir = os.path.join(dir_path)
    cellcomm_adr = get_column_letter(int(cellcomm[0])) + str(cellcomm[1])  # getting adresses of cells to write to
    cellvalue_adr = get_column_letter(int(cellvalue[0])) + str(cellvalue[1])
    reg_book = load_workbook(target_dir + '/' + files_reg, read_only=True)
    wrksheet_reg = reg_book.active
    os.mkdir(target_dir + '/result')
    for row in range(1, wrksheet_reg.max_row + 1):  # for rows in registry copies the values in new files
        tofill = load_workbook(target_dir + '/' + files_tofill, read_only=False)
        wrksheet_tofill = tofill.active
        comment = wrksheet_reg.cell(row=row, column=int(reg_com)).value
        value = wrksheet_reg.cell(row=row, column=int(reg_value)).value
        wrksheet_tofill[cellcomm_adr] = comment
        wrksheet_tofill[cellvalue_adr] = value
        tofill.save(target_dir + '/result/' + '{0}.xls'.format(comment))
        tofill.close()
    reg_book.close()
    res_arc = zipfile.ZipFile(target_dir + '/result/' + 'result.zip', 'w')  # ziping and deleting all new files
    for file in os.listdir(target_dir + '/result'):
        if file != 'result.zip':
            res_arc.write(target_dir + '/result/' + file)
    res_arc.close()


def form2Unit(dir_path, reg_com, reg_value):
    '''opens the regisrty and copies values to template files by rows'''

    target_dir = os.path.join(dir_path)
    file_list = os.listdir(target_dir)
    res_wb = Workbook()  # creating a workbook to write result registry
    res_wsht = res_wb.active
    res_wsht.title = 'Лист с результатами'
    target_column = 2  # count of number of column, where a new registry values will be added
    dict = {}
    for file in file_list:  # opening files source files and searching needed values
        wrkbook = load_workbook(target_dir + '/' + file, read_only=True)
        wrksheet = wrkbook.active

        for row in range(1, wrksheet.max_row + 1):
            cell_comm = wrksheet.cell(column=int(reg_com), row=row).value
            cell_val = wrksheet.cell(column=int(reg_value), row=row).value
            if cell_comm not in dict:
                dict[cell_comm] = {}
                dict[cell_comm][target_column] = cell_val
            else:
                dict[cell_comm][target_column] = cell_val
        target_column += 1
        wrkbook.close()
    for comms in dict.keys():
        inner_dict = dict[comms]
        i = 2
        res_list = []
        res_list.append(comms)

        for i in range(2, target_column):
            if i in inner_dict:
                res_list.append(inner_dict[i])
            else:
                res_list.append('')
        res_wsht.append(res_list)
    os.mkdir(target_dir + '/result')
    res_wb.save(target_dir + '/result/' + 'result.xlsx')
    res_wb.close()
